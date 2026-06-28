from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

wb = load_workbook("RetailIQ_Sensitivity_Model.xlsx")

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=16, bold=True, color="1B263B")
SUBTITLE = Font(name="Arial", size=11, italic=True, color="415A77")
HEADER_FILL = PatternFill("solid", start_color="1B263B")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SCENARIO_FILL = {
    "conservative": PatternFill("solid", start_color="FCE4E4"),
    "moderate": PatternFill("solid", start_color="FFF4D6"),
    "optimistic": PatternFill("solid", start_color="E2F0D9"),
    "custom": PatternFill("solid", start_color="DCE6F1"),
}

def style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

ws2 = wb.create_sheet("Sensitivity Model")
ws2["A1"] = "RetailIQ — Sensitivity Analysis: Revenue Protected Under 3 Scenarios"
ws2["A1"].font = TITLE
ws2["A2"] = "Green cells link to Assumptions sheet. Black cells are calculated formulas — change any blue input on Assumptions to see this recalculate."
ws2["A2"].font = SUBTITLE

states = ["AL", "MA", "PI", "CE"]
assumption_rows = {"AL": 6, "MA": 7, "PI": 8, "CE": 9}

def build_scenario_block(ws, start_row, title, fill_key, target_formula_fn):
    ws.cell(row=start_row, column=1, value=title).font = BOLD
    headers = ["State", "Total Orders", "Late Orders", "Current Late %", "Target Late %", "Orders Shifted", "Revenue Protected (R$)"]
    hrow = start_row + 1
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=i, value=h)
    style_header_row(ws, hrow, 1, len(headers))

    for i, state in enumerate(states):
        row = hrow + 1 + i
        a_row = assumption_rows[state]
        ws.cell(row=row, column=1, value=f"=Assumptions!A{a_row}").font = GREEN
        ws.cell(row=row, column=2, value=f"=Assumptions!B{a_row}").font = GREEN
        ws.cell(row=row, column=3, value=f"=Assumptions!C{a_row}").font = GREEN
        ws.cell(row=row, column=4, value=f"=Assumptions!D{a_row}").font = GREEN
        ws.cell(row=row, column=4).number_format = "0.0%"

        target_formula = target_formula_fn(row, a_row)
        ws.cell(row=row, column=5, value=target_formula).font = BLACK
        ws.cell(row=row, column=5).number_format = "0.0%"

        ws.cell(row=row, column=6, value=f"=MAX(C{row}-B{row}*E{row},0)").font = BLACK
        ws.cell(row=row, column=6).number_format = "0"

        ws.cell(row=row, column=7, value=f"=F{row}*Assumptions!$F$10").font = BLACK
        ws.cell(row=row, column=7).number_format = "R$#,##0.00"

        for c in range(1, 8):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).fill = SCENARIO_FILL[fill_key]

    total_row = hrow + 1 + len(states)
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD
    ws.cell(row=total_row, column=6, value=f"=SUM(F{hrow+1}:F{total_row-1})").font = BOLD
    ws.cell(row=total_row, column=6).number_format = "0"
    ws.cell(row=total_row, column=7, value=f"=SUM(G{hrow+1}:G{total_row-1})").font = BOLD
    ws.cell(row=total_row, column=7).number_format = "R$#,##0.00"
    for c in range(1, 8):
        ws.cell(row=total_row, column=c).border = BORDER
        ws.cell(row=total_row, column=c).fill = PatternFill("solid", start_color="D9D9D9")

    return total_row

# Scenario 1: Conservative (each state -5pp from its own baseline)
r1 = build_scenario_block(
    ws2, 4, "Scenario 1: Conservative — each state improves by 5 percentage points", "conservative",
    lambda row, a_row: f"=MAX(D{row}-Assumptions!$B$17/100,0)"
)

# Scenario 2: Moderate (all states to national average)
r2 = build_scenario_block(
    ws2, r1 + 3, "Scenario 2: Moderate — all 4 states reach the national average late rate", "moderate",
    lambda row, a_row: "=Assumptions!$B$13"
)

# Scenario 3: Optimistic (all states to best-performing state level)
r3 = build_scenario_block(
    ws2, r2 + 3, "Scenario 3: Optimistic — all 4 states match the best-performing state (Roraima)", "optimistic",
    lambda row, a_row: "=Assumptions!$B$14"
)

# Custom interactive scenario, fully driven by Assumptions!B17
r4 = build_scenario_block(
    ws2, r3 + 3, "Custom Scenario — driven by the yellow input cell on the Assumptions sheet", "custom",
    lambda row, a_row: f"=MAX(D{row}-Assumptions!$B$17/100,0)"
)

col_widths = {"A": 10, "B": 14, "C": 13, "D": 15, "E": 14, "F": 16, "G": 20}
for col, width in col_widths.items():
    ws2.column_dimensions[col].width = width

wb.save("RetailIQ_Sensitivity_Model.xlsx")
print(f"Sheet 2 built. Scenario blocks end at rows: {r1}, {r2}, {r3}, {r4}")
