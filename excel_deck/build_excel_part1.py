"""
RetailIQ — MBB-style Sensitivity Model
Real numbers from sql/02_analysis.sql Q4 (state delivery performance) and
python/sensitivity_analysis.py. All scenario math is live Excel formulas,
not hardcoded Python-computed values.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

wb = Workbook()

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=16, bold=True, color="1B263B")
SUBTITLE = Font(name="Arial", size=11, italic=True, color="415A77")
HEADER_FILL = PatternFill("solid", start_color="1B263B")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
YELLOW = PatternFill("solid", start_color="FFFF00")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

# ============================================================
# SHEET 1: Assumptions
# ============================================================
ws1 = wb.active
ws1.title = "Assumptions"
ws1["A1"] = "RetailIQ — Sensitivity Model: Assumptions & Inputs"
ws1["A1"].font = TITLE
ws1["A2"] = "All blue cells are hardcoded inputs sourced from real Olist data. Black cells are formulas."
ws1["A2"].font = SUBTITLE

ws1["A4"] = "State-Level Delivery Performance (4 worst-performing states)"
ws1["A4"].font = BOLD

headers = ["State", "Total Orders", "Late Orders", "Current Late %", "Total Revenue (R$)", "Avg Order Value (R$)"]
for i, h in enumerate(headers, start=1):
    ws1.cell(row=5, column=i, value=h)
style_header_row(ws1, 5, 1, len(headers))

state_data = [
    ("AL", 427, 103, 94172.49),
    ("MA", 800, 163, 147803.55),
    ("PI", 523, 81, 105178.19),
    ("CE", 1426, 218, 266436.77),
]
for i, (state, orders, late, revenue) in enumerate(state_data):
    row = 6 + i
    ws1.cell(row=row, column=1, value=state).font = BLUE
    ws1.cell(row=row, column=2, value=orders).font = BLUE
    ws1.cell(row=row, column=3, value=late).font = BLUE
    ws1.cell(row=row, column=4, value=f"=C{row}/B{row}").font = BLACK
    ws1.cell(row=row, column=4).number_format = "0.0%"
    ws1.cell(row=row, column=5, value=revenue).font = BLUE
    ws1.cell(row=row, column=6, value=f"=E{row}/B{row}").font = BLACK
    ws1.cell(row=row, column=6).number_format = "R$#,##0.00"
    for c in range(1, 7):
        ws1.cell(row=row, column=c).border = BORDER

# Totals row
total_row = 10
ws1.cell(row=total_row, column=1, value="TOTAL / BLENDED").font = BOLD
ws1.cell(row=total_row, column=2, value="=SUM(B6:B9)").font = BLACK
ws1.cell(row=total_row, column=3, value="=SUM(C6:C9)").font = BLACK
ws1.cell(row=total_row, column=4, value=f"=C{total_row}/B{total_row}").font = BLACK
ws1.cell(row=total_row, column=4).number_format = "0.0%"
ws1.cell(row=total_row, column=5, value="=SUM(E6:E9)").font = BLACK
ws1.cell(row=total_row, column=5).number_format = "R$#,##0.00"
ws1.cell(row=total_row, column=6, value=f"=E{total_row}/B{total_row}").font = BLACK
ws1.cell(row=total_row, column=6).number_format = "R$#,##0.00"
for c in range(1, 7):
    ws1.cell(row=total_row, column=c).border = BORDER
    ws1.cell(row=total_row, column=c).fill = PatternFill("solid", start_color="E8EEF4")

ws1.cell(row=5, column=2).comment = Comment(
    "Source: RetailIQ sql/02_analysis.sql Q4 (delivery performance by state), "
    "computed from Olist Brazilian E-Commerce dataset, orders with status='delivered'.",
    "RetailIQ Analysis"
)

ws1["A12"] = "Benchmark Late-Delivery Rates"
ws1["A12"].font = BOLD
ws1["A13"] = "National average late-delivery rate"
ws1["B13"] = 0.0811
ws1["B13"].font = BLUE
ws1["B13"].number_format = "0.00%"
ws1["B13"].comment = Comment("Source: sql/02_analysis.sql — national avg across all delivered orders, n=96,470.", "RetailIQ Analysis")
ws1["A14"] = "Best-performing state (Roraima/RO) late rate"
ws1["B14"] = 0.029
ws1["B14"].font = BLUE
ws1["B14"].number_format = "0.00%"
ws1["B14"].comment = Comment("Source: sql/02_analysis.sql Q4 full state ranking — RO is the best-performing state nationally.", "RetailIQ Analysis")

ws1["A16"] = "Custom Scenario Input (change this to test any improvement level)"
ws1["A16"].font = BOLD
ws1["A17"] = "Percentage-point reduction in late rate (per state)"
ws1["B17"] = 5
ws1["B17"].font = BOLD
ws1["B17"].fill = YELLOW
ws1["B17"].number_format = "0.0"
ws1["B17"].border = BORDER
ws1["C17"] = "← Change this number; the Sensitivity Model sheet recalculates automatically"
ws1["C17"].font = Font(name="Arial", size=9, italic=True, color="808080")

col_widths = {"A": 28, "B": 16, "C": 16, "D": 16, "E": 20, "F": 20}
for col, width in col_widths.items():
    ws1.column_dimensions[col].width = width

print("Sheet 1 (Assumptions) built")
wb.save("RetailIQ_Sensitivity_Model.xlsx")
