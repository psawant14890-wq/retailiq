from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook("RetailIQ_Sensitivity_Model.xlsx")

TITLE = Font(name="Arial", size=18, bold=True, color="1B263B")
SUBTITLE = Font(name="Arial", size=11, italic=True, color="415A77")
BOLD = Font(name="Arial", size=10, bold=True)
GREEN = Font(name="Arial", size=10, color="008000")
BLACK = Font(name="Arial", size=10, color="000000")
KPI_LABEL = Font(name="Arial", size=10, color="595959")
KPI_VALUE = Font(name="Arial", size=20, bold=True, color="1B263B")
KPI_VALUE_RISK = Font(name="Arial", size=20, bold=True, color="C00000")
SECTION = Font(name="Arial", size=13, bold=True, color="1B263B")
BODY = Font(name="Arial", size=10, color="000000")
KPI_FILL = PatternFill("solid", start_color="F0F4F8")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ws3 = wb.create_sheet("Executive Summary", 0)  # insert as first sheet

ws3["A1"] = "RetailIQ — Executive Summary"
ws3["A1"].font = TITLE
ws3["A2"] = "E-Commerce Revenue & Delivery Performance Analytics | Olist Brazilian E-Commerce Dataset"
ws3["A2"].font = SUBTITLE

# Situation / Complication / Recommendation
ws3["A4"] = "Situation"
ws3["A4"].font = SECTION
ws3["A5"] = ("Olist is a multi-seller e-commerce marketplace connecting ~3,000 small sellers to customers "
             "across Brazil. Revenue grew from R$280 in Sept 2016 to a peak of over R$1.17M/month by Nov 2017.")
ws3["A5"].font = BODY
ws3["A5"].alignment = Alignment(wrap_text=True)
ws3.merge_cells("A5:H5")
ws3.row_dimensions[5].height = 30

ws3["A7"] = "Complication"
ws3["A7"].font = SECTION
ws3["A8"] = ("Delivery delays are strongly associated with customer dissatisfaction. Orders delivered on time "
             "average a 4.28/5 review score; orders delivered 7+ days late average just 1.73/5. R$607,588 in "
             "revenue is tied to orders in this severe-delay bucket, concentrated disproportionately in 4 states.")
ws3["A8"].font = BODY
ws3["A8"].alignment = Alignment(wrap_text=True)
ws3.merge_cells("A8:H8")
ws3.row_dimensions[8].height = 45

ws3["A10"] = "Recommendation"
ws3["A10"].font = SECTION
ws3["A11"] = ("Prioritize logistics intervention in AL, MA, PI, and CE (the 4 worst-performing states, all >15% "
              "late-delivery rate) rather than a blanket logistics overhaul. Even the conservative scenario "
              "(a 5-point improvement) protects ~R$30,700 in revenue — the recommendation holds under a "
              "pessimistic assumption, not just the best case.")
ws3["A11"].font = BODY
ws3["A11"].alignment = Alignment(wrap_text=True)
ws3.merge_cells("A11:H11")
ws3.row_dimensions[11].height = 45

# KPI cards row
ws3["A13"] = "Key Metrics"
ws3["A13"].font = SECTION

kpis = [
    ("Total Revenue", "R$ 15.84M", "B"),
    ("Late Delivery Rate", "8.1%", "D"),
    ("Avg Review Score", "4.07 / 5", "F"),
    ("Revenue at Risk (7+ Days Late)", "R$ 607.6K", "H"),
]
kpi_row = 14
for label, value, col in kpis:
    col_idx = ws3[f"{col}1"].column
    ws3.cell(row=kpi_row, column=col_idx, value=label).font = KPI_LABEL
    ws3.cell(row=kpi_row + 1, column=col_idx, value=value).font = (
        KPI_VALUE_RISK if "Risk" in label else KPI_VALUE
    )
    for r in (kpi_row, kpi_row + 1):
        ws3.cell(row=r, column=col_idx).fill = KPI_FILL
        ws3.cell(row=r, column=col_idx).border = BORDER

ws3.row_dimensions[15].height = 28

# Sensitivity summary table, linked live to Sensitivity Model sheet
ws3["A18"] = "Sensitivity Summary — Revenue Protected by Scenario"
ws3["A18"].font = SECTION

headers = ["Scenario", "Target Approach", "Orders Shifted to On-Time", "Revenue Protected (R$)"]
for i, h in enumerate(headers, start=1):
    cell = ws3.cell(row=19, column=i, value=h)
    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="1B263B")
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center")

scenario_links = [
    ("Conservative", "Each state improves 5pp", "='Sensitivity Model'!F10", "='Sensitivity Model'!G10"),
    ("Moderate", "All 4 states reach national avg (8.11%)", "='Sensitivity Model'!F19", "='Sensitivity Model'!G19"),
    ("Optimistic", "All 4 states match best state (2.9%)", "='Sensitivity Model'!F28", "='Sensitivity Model'!G28"),
]
for i, (name, desc, orders_f, revenue_f) in enumerate(scenario_links):
    row = 20 + i
    ws3.cell(row=row, column=1, value=name).font = BOLD
    ws3.cell(row=row, column=2, value=desc).font = BODY
    ws3.cell(row=row, column=3, value=orders_f).font = GREEN
    ws3.cell(row=row, column=4, value=revenue_f).font = GREEN
    ws3.cell(row=row, column=4).number_format = "R$#,##0.00"
    for c in range(1, 5):
        ws3.cell(row=row, column=c).border = BORDER

ws3["A24"] = "Note: this table updates automatically if you change the yellow input cell on the Assumptions sheet."
ws3["A24"].font = Font(name="Arial", size=9, italic=True, color="808080")

col_widths = {"A": 18, "B": 30, "C": 22, "D": 20, "E": 4, "F": 18, "G": 4, "H": 20}
for col, width in col_widths.items():
    ws3.column_dimensions[col].width = width

wb.save("RetailIQ_Sensitivity_Model.xlsx")
print("Sheet 3 (Executive Summary) built and inserted as first sheet")
